# -*- coding: utf-8 -*-
"""
CAN_TestFrame_TOOL (PyQt5) V1.0.0
功能：
  1. 顶部菜单栏 文件 / 配置 / 帮助；文件下拉含 导入 子菜单（IO枚举文本、Debug表格）；
  2. 刚打开仅显示菜单栏，导入文件后才出现对应界面；
  3. IO枚举文本：导入 I/O 枚举 txt，识别 NUM_OF_INPUTS / NUM_OF_OUTPUTS，
     生成 app_can_test_frame_01_input_output.c 及上位机 ini/DBC 文件（Motorola 格式）；
  4. Debug表格：导入 Excel，按生成规则.md 解析各 Sheet CAN 信号，
     生成 CAN 测试帧 C 函数及每 Sheet 独立的 ini/DBC 文件。
  5. 帮助：使用教程（打开内嵌的 HTML 说明书）、联系我们；版本号在 配置 → 关于 弹窗中展示。
"""
import re
import sys
import os
import math
import shutil

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QIcon, QKeySequence
from PyQt5.QtWidgets import (
    QAction, QApplication, QDialog, QFileDialog, QHBoxLayout, QHeaderView,
    QLabel, QLineEdit, QMainWindow, QMessageBox, QPlainTextEdit, QPushButton,
    QSplitter, QTabWidget, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
    QAbstractItemView,
)

APP_NAME = 'CAN_TestFrame_TOOL'
APP_VERSION = 'V1.0.0'

IDENT_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')
SIG_NAME_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_\.\[\]\(\)\,\>\- ]*$')

MANUAL_HTML = 'CAN_TestFrame_TOOL使用说明书.html'
LOGO_PNG = 'logo_cat.png'


def resource_path(rel):
    """获取内嵌资源路径（兼容 PyInstaller onefile 打包）。"""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


def read_text_file(path):
    for enc in ('utf-8-sig', 'utf-8', 'gb18030', 'latin-1'):
        try:
            with open(path, 'r', encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        return f.read()


def strip_comments(line):
    line = line.split('//', 1)[0]
    line = re.sub(r'/\*.*?\*/', '', line)
    return line


def extract_enum_block(text, enum_name):
    close_m = re.search(r'\}\s*' + re.escape(enum_name) + r'\s*;', text)
    if not close_m:
        return None
    close_pos = close_m.start()
    prefix = text[:close_pos]
    type_matches = list(re.finditer(r'typedef\s+enum', prefix))
    if not type_matches:
        return None
    type_end = type_matches[-1].end()
    brace_m = re.search(r'\{', text[type_end:close_pos])
    if not brace_m:
        return None
    brace_pos = type_end + brace_m.start()
    return text[brace_pos + 1: close_pos]


def parse_members(text, enum_name, marker):
    block = extract_enum_block(text, enum_name)
    if block is None:
        return None, False
    members = []
    marker_found = False
    for raw in block.split('\n'):
        line = strip_comments(raw)
        if not line.strip():
            continue
        for tok in line.split(','):
            tok = tok.strip()
            if not tok:
                continue
            name = tok.split('=')[0].strip()
            if not IDENT_RE.match(name):
                continue
            if name == marker:
                marker_found = True
                break
            members.append(name)
        if marker_found:
            break
    return members, marker_found


def parse_io_enum(text):
    inputs, in_ok = parse_members(text, 'INPUTS_ENUM', 'NUM_OF_INPUTS')
    outputs, out_ok = parse_members(text, 'OUTPUTS_ENUM', 'NUM_OF_OUTPUTS')
    if inputs is None or outputs is None or not in_ok or not out_ok:
        raise ValueError(
            '未能识别 INPUTS_ENUM / OUTPUTS_ENUM 或 '
            'NUM_OF_INPUTS / NUM_OF_OUTPUTS 标记，请检查 txt 文件格式。')
    return {
        'num_inputs': len(inputs),
        'num_outputs': len(outputs),
        'inputs': inputs,
        'outputs': outputs,
    }


def generate_function_text(num_inputs, num_outputs):
    in_can_frames = (num_inputs + 63) // 64
    out_can_frames = (num_outputs + 63) // 64
    out_start = in_can_frames * 8
    total_buf = out_start + out_can_frames * 8

    L = []
    L.append('static void app_can_test_frame_01_input_output(void)')
    L.append('{')
    L.append('    UINT8 u8_index;')
    L.append('')

    in_full = num_inputs // 8
    in_rem = num_inputs % 8
    L.append('    //-input NUM_OF_INPUTS:%d(0~%d)--------------------------------------------------------------'
             % (num_inputs, num_inputs - 1))
    if in_full > 0 and in_rem > 0:
        L.append('    /* %d inputs: byte 0~%d (input 0~%d, %d full bytes) + byte %d (input %d~%d, %d bits) */'
                 % (num_inputs, in_full - 1, in_full * 8 - 1, in_full,
                    in_full, in_full * 8, num_inputs - 1, in_rem))
    elif in_full > 0:
        L.append('    /* %d inputs: byte 0~%d (input 0~%d, %d full bytes) */'
                 % (num_inputs, in_full - 1, num_inputs - 1, in_full))
    else:
        L.append('    /* %d inputs: byte 0 (input 0~%d, %d bits) */'
                 % (num_inputs, num_inputs - 1, in_rem))

    if in_full > 0:
        L.append('    for (u8_index = 0u; u8_index < %du; u8_index++) /* input 0~%d */'
                 % (in_full, in_full * 8 - 1))
        L.append('    {')
        for b in range(8):
            if b == 0:
                L.append('        g_u8_tx_can_data[u8_index]  =  HWA_GetInputDebState(u8_index * 8u);')
            else:
                L.append('        g_u8_tx_can_data[u8_index] |= (HWA_GetInputDebState(u8_index * 8u + %du) << %du);'
                         % (b, b))
        L.append('    }')
        L.append('')

    if in_rem > 0:
        L.append('    /* byte %d: input %d~%d (%d bits) */'
                 % (in_full, in_full * 8, num_inputs - 1, in_rem))
        for b in range(in_rem):
            idx = in_full * 8 + b
            if b == 0:
                L.append('    g_u8_tx_can_data[%d]  =  HWA_GetInputDebState(%du);' % (in_full, idx))
            else:
                L.append('    g_u8_tx_can_data[%d] |= (HWA_GetInputDebState(%du) << %du);' % (in_full, idx, b))
        L.append('')

    in_end = in_full + (1 if in_rem else 0)
    if in_end < out_start:
        L.append('    /* padding: bytes %d~%d (align to next CAN frame) */' % (in_end, out_start - 1))
        for i in range(in_end, out_start):
            L.append('    g_u8_tx_can_data[%d] = 0u;' % i)
        L.append('')

    out_full = num_outputs // 8
    out_rem = num_outputs % 8
    L.append('    //-output NUM_OF_OUTPUTS:%d(0~%d)------------------------------------------------------------'
             % (num_outputs, num_outputs - 1))
    if out_full > 0 and out_rem > 0:
        L.append('    /* %d outputs: byte %d~%d (output 0~%d, %d full bytes) + byte %d (output %d~%d, %d bits) */'
                 % (num_outputs, out_start, out_start + out_full - 1, out_full * 8 - 1, out_full,
                    out_start + out_full, out_full * 8, num_outputs - 1, out_rem))
    elif out_full > 0:
        L.append('    /* %d outputs: byte %d~%d (output 0~%d, %d full bytes) */'
                 % (num_outputs, out_start, out_start + out_full - 1, num_outputs - 1, out_full))
    else:
        L.append('    /* %d outputs: byte %d (output 0~%d, %d bits) */'
                 % (num_outputs, out_start, num_outputs - 1, out_rem))

    if out_full > 0:
        L.append('    for (u8_index = %du; u8_index < %du; u8_index++) /* output 0~%d */'
                 % (out_start, out_start + out_full, out_full * 8 - 1))
        L.append('    {')
        for b in range(8):
            if b == 0:
                L.append('        if (HWA_GetOutputStatus((u8_index - %du) * 8u) > 0u)' % out_start)
                L.append('        {')
                L.append('            g_u8_tx_can_data[u8_index] = BIT0;')
                L.append('        }')
                L.append('        else')
                L.append('        {')
                L.append('            g_u8_tx_can_data[u8_index] = 0u;')
                L.append('        }')
            else:
                L.append('        if (HWA_GetOutputStatus((u8_index - %du) * 8u + %du) > 0u)' % (out_start, b))
                L.append('        {')
                L.append('            g_u8_tx_can_data[u8_index] |= BIT%d;' % b)
                L.append('        }')
        L.append('    }')
        L.append('')

    if out_rem > 0:
        L.append('    /* byte %d: output %d~%d (%d bits) */'
                 % (out_start + out_full, out_full * 8, num_outputs - 1, out_rem))
        L.append('    g_u8_tx_can_data[%d] = 0u;' % (out_start + out_full))
        for b in range(out_rem):
            idx = out_full * 8 + b
            L.append('    if (HWA_GetOutputStatus(%du) > 0u)' % idx)
            L.append('    {')
            L.append('        g_u8_tx_can_data[%d] |= BIT%d;' % (out_start + out_full, b))
            L.append('    }')
        L.append('')

    out_end = out_start + out_full + (1 if out_rem else 0)
    if out_end < total_buf:
        L.append('    /* padding: bytes %d~%d (align to next CAN frame) */' % (out_end, total_buf - 1))
        for i in range(out_end, total_buf):
            L.append('    g_u8_tx_can_data[%d] = 0u;' % i)
        L.append('')

    L.append('}')
    return '\n'.join(L)


def generate_full_c_content(num_inputs, num_outputs):
    header = (
        '/******************************************************************************\n'
        ' * FILE NAME:    app_can_test_frame_01_input_output.c\n'
        ' * DESCRIPTION:  Generated by %s %s\n'
        ' * NUM_OF_INPUTS  = %d\n'
        ' * NUM_OF_OUTPUTS = %d\n'
        ' *****************************************************************************/\n'
        '\n' % (APP_NAME, APP_VERSION, num_inputs, num_outputs))
    return header + generate_function_text(num_inputs, num_outputs) + '\n'


def _signal_start(j):
    """第 j 个 1-bit signal 的 start bit（顺序编号，与参考 ini 一致）。"""
    return j


def generate_ini_content(inputs, outputs, can_ids):
    """生成上位机 ini 文件，Motorola 格式，每位输入/输出作为 1-bit Signal。"""
    L = []
    L.append('///配置说明：')
    L.append('///每行以$xx yyy:起始的为有效行，$xx yyy:请勿修改')
    L.append("///每行$xx yyy:之后的为该服务的数据，数据中以','分开")
    L.append('///每行以///起始的为注释行，说明之后的数据存放格式，请勿删除')
    L.append('///注意：每行除空白行、注释行和有效行之外任何数据均会报错')
    L.append('')
    L.append("///$Message: + Id + ','  + Length + ','  + Name + ','  + Cycle")
    L.append('///&Format: + Intel or Motorola or Lin')
    L.append("///@Signal: + Start + ','  + Length + ','  + Name + ',' + DefaultValue")
    L.append("///#Value: + Value + ','  + Name")
    L.append("///#PhyValue: + MaxValue + ','  + Factor + ',' + Offset + ',' + Unit")
    L.append('')

    idx = 0
    for i in range(0, len(inputs), 64):
        chunk = inputs[i:i + 64]
        msg_num = i // 64 + 1
        msg_id = can_ids[idx]
        idx += 1
        L.append('$Message:%s,8,输入状态%d,10' % (msg_id, msg_num))
        L.append('&Format:Motorola')
        for j, name in enumerate(chunk):
            L.append('@Signal:%d,1,%s,0' % (_signal_start(j), name))
            L.append('#Value:0,OFF;')
            L.append('#Value:1,ON;')
        L.append('')

    for i in range(0, len(outputs), 64):
        chunk = outputs[i:i + 64]
        msg_num = i // 64 + 1
        msg_id = can_ids[idx]
        idx += 1
        L.append('$Message:%s,8,输出状态%d,10' % (msg_id, msg_num))
        L.append('&Format:Motorola')
        for j, name in enumerate(chunk):
            L.append('@Signal:%d,1,%s,0' % (_signal_start(j), name))
            L.append('#Value:0,OFF;')
            L.append('#Value:1,ON;')
        L.append('')

    return '\n'.join(L)


def generate_dbc_content(inputs, outputs, can_ids):
    """生成 DBC 文件内容，适配 Vector CANoe/CANalyzer。"""
    L = []
    L.append('VERSION " "')
    L.append('')
    L.append('')
    L.append('')
    L.append(' NS_ : ')
    for ns in ['NS_DESC_', 'CM_', 'BA_DEF_', 'BA_', 'VAL_', 'CAT_DEF_', 'CAT_',
               'FILTER', 'BA_DEF_DEF_', 'EV_DATA_', 'ENVVAR_DATA_', 'SGTYPE_',
               'SGTYPE_VAL_', 'BA_DEF_SGTYPE_', 'SIG_TYPE_REF_', 'SIG_GROUP_',
               'SIG_VALTYPE_', 'SIGTYPE_VALTYPE_', 'BO_TX_BU_', 'BA_DEF_REL_',
               'BA_REL_', 'BA_DEF_DEF_REL_', 'BU_SG_REL_', 'BU_EV_REL_',
               'BU_BO_REL_', 'SG_MUL_VAL_']:
        L.append('      ' + ns)
    L.append('')
    L.append('BS_:')
    L.append('')
    L.append('BU_: BCM')
    L.append('')
    L.append('')
    L.append('//monitor')
    L.append('')

    idx = 0
    msg_info = []
    for i in range(0, len(inputs), 64):
        chunk = inputs[i:i + 64]
        hex_id = can_ids[idx].upper()
        dec_id = int(hex_id, 16)
        msg_name = 'Tx_%s' % hex_id
        msg_info.append((dec_id, chunk, 'input'))
        L.append('BO_ %d %s: 8 BCM' % (dec_id, msg_name))
        for j, name in enumerate(chunk):
            L.append('   SG_ %s: %d|1@0+ (1,0) [0|0] ""  Vector__XXX' % (name, j))
        L.append('')
        idx += 1

    L.append('//simulation')
    L.append('')

    for i in range(0, len(outputs), 64):
        chunk = outputs[i:i + 64]
        hex_id = can_ids[idx].upper()
        dec_id = int(hex_id, 16)
        msg_name = 'Tx_%s' % hex_id
        msg_info.append((dec_id, chunk, 'output'))
        L.append('BO_ %d %s: 8 BCM' % (dec_id, msg_name))
        for j, name in enumerate(chunk):
            L.append('   SG_ %s: %d|1@0+ (1,0) [0|0] ""  Vector__XXX' % (name, j))
        L.append('')
        idx += 1

    L.append('BA_DEF_ BO_  "GenMsgSendType" ENUM  "Cyclic","Event";')
    L.append('BA_DEF_ BO_  "GenMsgCycleTime" INT 0 65535;')
    L.append('BA_DEF_  "BusType" STRING ;')
    L.append('BA_DEF_DEF_  "GenMsgSendType" "Cyclic";')
    L.append('BA_DEF_DEF_  "GenMsgCycleTime" 0;')
    L.append('BA_DEF_DEF_  "BusType" "CAN";')
    L.append('BA_ "BusType" "CAN1";')
    L.append('')
    L.append('')

    for dec_id, _, _ in msg_info:
        L.append('BA_ "GenMsgCycleTime" BO_ %d 10;' % dec_id)
    L.append('')

    for dec_id, chunk, _ in msg_info:
        for name in chunk:
            L.append('VAL_ %d %s 0 "OFF" 1 "ON";' % (dec_id, name))

    return '\n'.join(L)


# ============================================================ Debug_Excel

def _sanitize_c_name(name):
    return re.sub(r'[^A-Za-z0-9_]', '_', name)


def _get_merge_rows(ws, row, col):
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= row <= rng.max_row and
                rng.min_col <= col <= rng.max_col):
            return rng.max_row - rng.min_row + 1
    return 1


def _is_merged_start(ws, row, col):
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= row <= rng.max_row and
                rng.min_col <= col <= rng.max_col):
            return rng.min_row == row
    return True


def _cell_val(ws, row, col):
    for rng in ws.merged_cells.ranges:
        if (rng.min_row <= row <= rng.max_row and
                rng.min_col <= col <= rng.max_col):
            v = ws.cell(row=rng.min_row, column=rng.min_col).value
            break
    else:
        v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def _check_sig_name(s):
    if re.search(r'[\u4e00-\u9fff\u3000-\u303f\uff00-\uffef]', s):
        return False
    if not SIG_NAME_RE.match(s):
        return False
    return True


def parse_excel_file(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = []
    errors = []
    warnings = []
    if len(wb.sheetnames) < 2:
        errors.append('Excel 需至少包含 2 个工作表（第 1 个工作表将被忽略），当前仅 %d 个。'
                      % len(wb.sheetnames))
    for idx, name in enumerate(wb.sheetnames):
        if idx == 0:
            continue
        ws = wb[name]
        hdr_row = None
        for r in range(1, min(ws.max_row, 20) + 1):
            a = _cell_val(ws, r, 1)
            if a is not None and str(a).strip().lower() in ('byte', 'bit'):
                hdr_row = r
                break
        if hdr_row is None:
            errors.append("工作表 '%s'：前 20 行内未找到参数表头行（A 列应为 Byte）。" % name)
            continue
        col_map = {}
        for c in range(1, ws.max_column + 1):
            h = _cell_val(ws, hdr_row, c)
            if h is None:
                continue
            hs = str(h).strip().lower()
            if hs in ('byte', 'bit') and c == 1:
                col_map['byte'] = c
            elif hs in ('bit', 'bit位') and c == 2:
                col_map['bit'] = c
            elif 'define' in hs or 'signal' in hs.lower():
                col_map['signal'] = c
            elif 'note' in hs.lower():
                col_map['note'] = c
            elif hs == 'id':
                col_map['can_id'] = c
        if 'byte' not in col_map:
            col_map['byte'] = 1
        if 'bit' not in col_map:
            col_map['bit'] = 2
        if 'signal' not in col_map:
            col_map['signal'] = 3
        if 'can_id' not in col_map:
            for c in range(4, ws.max_column + 1):
                h = _cell_val(ws, hdr_row, c)
                if h and str(h).strip().upper() == 'ID':
                    col_map['can_id'] = c
                    break
        data_start = hdr_row + 1
        if 'can_id' not in col_map:
            hex_re = re.compile(r'^(0[xX])?[0-9A-Fa-f]{2,3}$')
            for c in range(4, ws.max_column + 1):
                hits = 0
                for r in range(data_start, min(data_start + 20, ws.max_row + 1)):
                    v = _cell_val(ws, r, c)
                    if v is not None and hex_re.match(str(v).strip()):
                        hits += 1
                if hits >= 2:
                    col_map['can_id'] = c
                    if col_map.get('note') == c:
                        del col_map['note']
                    break
        # 数据区内合并形态校验 + 索引构建（表头行及以上的装饰性合并不校验）
        a_merge = {}
        c_merge = {}
        for rng in ws.merged_cells.ranges:
            if rng.max_row <= hdr_row:
                continue
            # 3.4/3.5/3.6：A/B/C 只允许纵向单列合并，B 列数据区内禁止合并
            for col_idx in (1, 2, 3):
                if rng.min_col <= col_idx <= rng.max_col:
                    col_name = {1: 'A', 2: 'B', 3: 'C'}[col_idx]
                    if rng.min_col != rng.max_col:
                        errors.append("工作表 '%s' 第 %d~%d 行：数据区内 %s 列不允许横向跨列合并（%s）。"
                                      % (name, rng.min_row, rng.max_row, col_name, str(rng)))
                    elif col_idx == 2:
                        errors.append("工作表 '%s' 第 %d~%d 行：B 列（Bit 列）不得使用合并单元格。"
                                      % (name, rng.min_row, rng.max_row))
            if rng.min_col == rng.max_col == 1:
                for rr in range(rng.min_row, rng.max_row + 1):
                    a_merge[rr] = (rng.min_row, rng.max_row)
            elif rng.min_col == rng.max_col == 3 and rng.max_row > rng.min_row:
                c_merge[rng.min_row] = rng.max_row

        # 3.6 约束 2/3：C 合并不得跨 byte 组；A 单格行上的 C 合并，其 A 列须为单格且 Byte 连续递增
        for c0, c1 in c_merge.items():
            m = a_merge.get(c0)
            if m is not None:
                if c1 > m[1]:
                    errors.append("工作表 '%s' 第 %d~%d 行：C 列合并不得跨越所在 byte 组边界（组为第 %d~%d 行）。"
                                  % (name, c0, c1, m[0], m[1]))
            else:
                span = c1 - c0 + 1
                byte_list = []
                for rr in range(c0, c1 + 1):
                    if a_merge.get(rr) is not None:
                        errors.append("工作表 '%s' 第 %d~%d 行：A 单格行上的 C 列合并范围内，第 %d 行 A 列应为单格。"
                                      % (name, c0, c1, rr))
                        byte_list = None
                        break
                    v = _cell_val(ws, rr, 1)
                    try:
                        byte_list.append(int(v))
                    except (ValueError, TypeError):
                        errors.append("工作表 '%s' 第 %d 行：C 列合并范围内 A 列 Byte 值 '%s' 不是整数。"
                                      % (name, rr, v))
                        byte_list = None
                        break
                if byte_list is not None:
                    if byte_list != list(range(byte_list[0], byte_list[0] + span)):
                        errors.append("工作表 '%s' 第 %d~%d 行：多字节信号（占 %d 字节）的 A 列 Byte 值 %s 必须连续递增。"
                                      % (name, c0, c1, span, byte_list))

        signals = []
        expected_byte = [0]

        def _check_byte_seq(val, row_idx):
            if val == 0:
                if expected_byte[0] != 0:
                    missing = ('byte %d' % expected_byte[0] if expected_byte[0] == 7
                               else 'byte %d~7' % expected_byte[0])
                    errors.append("工作表 '%s' 第 %d 行：前一组 Byte 止于 %d，未完整覆盖 0~7（缺少 %s）。"
                                  % (name, row_idx, expected_byte[0] - 1, missing))
                expected_byte[0] = 1
            elif val == expected_byte[0]:
                expected_byte[0] = (val + 1) % 8
            else:
                errors.append("工作表 '%s' 第 %d 行：Byte 值 %s 不符合 0~7 一组一组连续出现的规则（期望 %s）。"
                              % (name, row_idx, val, expected_byte[0]))
                expected_byte[0] = (val + 1) % 8

        def _emit_signal(row_idx, byte_val, bit_val):
            sig_val = _cell_val(ws, row_idx, col_map['signal'])
            if sig_val is None:
                return
            sig_val = str(sig_val).strip()
            if not sig_val:
                return
            if not _is_merged_start(ws, row_idx, col_map['signal']):
                return
            if not _check_sig_name(sig_val):
                errors.append("工作表 '%s' 第 %d 行：信号名 '%s' 不是合法的 C 表达式"
                              "（应以字母/下划线开头，不得包含中文或非法字符）。"
                              % (name, row_idx, sig_val))
                return
            note_val = _cell_val(ws, row_idx, col_map.get('note', 5))
            can_id_val = _cell_val(ws, row_idx, col_map.get('can_id', 6))
            merge_count = _get_merge_rows(ws, row_idx, col_map['signal'])
            if bit_val is not None:
                try:
                    bit_start = int(bit_val)
                except (ValueError, TypeError):
                    bit_start = 0
                bit_width = merge_count if merge_count > 1 else 1
            else:
                bit_start = 0
                bit_width = merge_count * 8 if merge_count > 1 else 8
            signals.append({
                'byte': byte_val,
                'bit_start': bit_start,
                'name': sig_val,
                'bit_width': bit_width,
                'note': str(note_val) if note_val else '',
                'can_id': str(can_id_val).strip() if can_id_val else '',
                'row': row_idx,
            })

        data_ended = False
        r = data_start
        while r <= ws.max_row:
            if data_ended:
                a_raw = _cell_val(ws, r, col_map['byte'])
                m = a_merge.get(r)
                if a_raw is not None or (m is not None and m[0] == r):
                    errors.append("工作表 '%s' 第 %d 行：Byte 列为空表示数据已结束，其后不应再出现 Byte 数据。"
                                  % (name, r))
                    break
                r += 1
                continue
            m = a_merge.get(r)
            if m is not None:
                if r != m[0]:
                    r += 1
                    continue
                span = m[1] - m[0] + 1
                if span != 8:
                    errors.append("工作表 '%s' 第 %d~%d 行：A 列 Byte 合并单元格应为 8 个（当前 %d 个）；"
                                  "单格表示整字节，8 格合并表示该字节含多位数据。"
                                  % (name, m[0], m[1], span))
                    r = m[1] + 1
                    continue
                a_raw = _cell_val(ws, m[0], 1)
                try:
                    byte_val = int(a_raw)
                except (ValueError, TypeError):
                    errors.append("工作表 '%s' 第 %d 行：A 列 Byte 值 '%s' 不是 0~7 的整数。"
                                  % (name, m[0], a_raw))
                    r = m[1] + 1
                    continue
                if not (0 <= byte_val <= 7):
                    errors.append("工作表 '%s' 第 %d 行：A 列 Byte 值 %s 超出 0~7 范围。"
                                  % (name, m[0], byte_val))
                    r = m[1] + 1
                    continue
                _check_byte_seq(byte_val, m[0])
                bvals = [_cell_val(ws, rr, col_map['bit']) for rr in range(m[0], m[1] + 1)]
                bad_b = None
                for bi, bv in enumerate(bvals):
                    try:
                        if int(bv) != bi:
                            bad_b = (bi, bv)
                            break
                    except (ValueError, TypeError):
                        bad_b = (bi, bv)
                        break
                if bad_b is not None:
                    errors.append("工作表 '%s' 第 %d~%d 行：8 格合并组内 B 列必须依次填写 0~7，"
                                  "第 %d 行 B 值 '%s' 不符合（期望 %d）。"
                                  % (name, m[0], m[1], m[0] + bad_b[0], bad_b[1], bad_b[0]))
                for rr in range(m[0], m[1] + 1):
                    _emit_signal(rr, byte_val, bvals[rr - m[0]])
                r = m[1] + 1
            else:
                a_raw = _cell_val(ws, r, col_map['byte'])
                if a_raw is None:
                    b_raw = _cell_val(ws, r, col_map['bit'])
                    c_raw = _cell_val(ws, r, col_map['signal'])
                    if b_raw is not None or c_raw is not None:
                        errors.append("工作表 '%s' 第 %d 行：Byte 列为空表示数据已结束，但该行 Bit/Signal 列仍有数据，"
                                      "请检查是否漏填 Byte 列。" % (name, r))
                        break
                    data_ended = True
                    continue
                try:
                    byte_val = int(a_raw)
                except (ValueError, TypeError):
                    errors.append("工作表 '%s' 第 %d 行：A 列 Byte 值 '%s' 不是 0~7 的整数。"
                                  % (name, r, a_raw))
                    r += 1
                    continue
                if not (0 <= byte_val <= 7):
                    errors.append("工作表 '%s' 第 %d 行：A 列 Byte 值 %s 超出 0~7 范围。"
                                  % (name, r, byte_val))
                    r += 1
                    continue
                _check_byte_seq(byte_val, r)
                bit_val = _cell_val(ws, r, col_map['bit'])
                if bit_val is not None:
                    errors.append("工作表 '%s' 第 %d 行：A 列为单个单元格时该行占用完整 1 字节，Bit 列应为空。"
                                  % (name, r))
                    r += 1
                    continue
                _emit_signal(r, byte_val, None)
                r += 1
        if expected_byte[0] != 0:
            missing = ('byte %d' % expected_byte[0] if expected_byte[0] == 7
                       else 'byte %d~7' % expected_byte[0])
            errors.append("工作表 '%s'：最后一组 Byte 止于 %d，未完整覆盖 0~7（缺少 %s）。"
                          % (name, expected_byte[0] - 1, missing))
        # 3.6 约束 5：同一帧（byte 0~7 组）内信号 bit 区间重叠告警；
        # byte 回到 0 表示新的一帧开始，重叠检测按帧独立进行
        occ_bits = {}
        prev_byte = None
        for sig in signals:
            if sig['byte'] == 0 and prev_byte is not None:
                occ_bits = {}
            g_start = sig['byte'] * 8 + sig['bit_start']
            g_end = g_start + sig['bit_width'] - 1
            clash = None
            for gb in range(g_start, g_end + 1):
                if gb in occ_bits and occ_bits[gb] != sig['name']:
                    clash = (gb, occ_bits[gb])
                    break
                occ_bits[gb] = sig['name']
            if clash is not None:
                warnings.append("工作表 '%s' 第 %d 行：信号 '%s' 与 '%s' 的 bit 区间重叠（byte %d bit %d），"
                                "后解析行将覆盖前面数据。"
                                % (name, sig['row'], sig['name'], clash[1],
                                   clash[0] // 8, clash[0] % 8))
            prev_byte = sig['byte']
        if signals:
            sheets.append({
                'name': name,
                'func_name': 'app_can_test_frame_' + _sanitize_c_name(name),
                'signals': signals,
            })
        else:
            errors.append("工作表 '%s'：表头行（第 %d 行）之后没有任何有效信号数据。" % (name, hdr_row))
    return sheets, errors, warnings


def _group_signals_by_byte(signals):
    groups = {}
    for sig in signals:
        b = sig['byte']
        if b not in groups:
            groups[b] = []
        groups[b].append(sig)
    return groups


def _get_can_id_list(signals):
    ids = []
    for sig in signals:
        cid = sig.get('can_id', '')
        if cid and cid not in ids:
            ids.append(cid)
    return ids if ids else ['']


def _determine_byte_strategy(byte_signals):
    if len(byte_signals) == 1 and byte_signals[0]['bit_width'] == 8:
        return 'direct'
    if all(s['bit_width'] == 1 for s in byte_signals):
        return 'bit'
    return 'shift'


def generate_excel_c_function(sheet_info):
    signals = sheet_info['signals']
    func_name = sheet_info['func_name']
    can_id_order = []
    can_id_map = {}
    for sig in signals:
        cid = sig.get('can_id', '')
        if cid and cid not in can_id_map:
            can_id_map[cid] = len(can_id_order)
            can_id_order.append(cid)
    if not can_id_order:
        can_id_order = ['']
        can_id_map[''] = 0
    total_bytes = len(can_id_order) * 8
    L = []
    L.append('static void %s(void)' % func_name)
    L.append('{')
    by_can = {}
    for sig in signals:
        cid = sig.get('can_id', '')
        if cid not in by_can:
            by_can[cid] = []
        by_can[cid].append(sig)
    defined_bytes = set()
    for cid_idx, cid in enumerate(can_id_order):
        cid_sigs = by_can.get(cid, [])
        byte_groups = _group_signals_by_byte(cid_sigs)
        for byte_idx in range(8):
            global_byte = cid_idx * 8 + byte_idx
            if global_byte in defined_bytes:
                continue
            if byte_idx in byte_groups:
                sigs = byte_groups[byte_idx]
                strategy = _determine_byte_strategy(sigs)
                defined_bytes.add(global_byte)
                if strategy == 'direct':
                    sig = sigs[0]
                    L.append('    g_u8_tx_can_data[%d] = %s;' % (global_byte, sig['name']))
                elif strategy == 'bit':
                    for i, sig in enumerate(sigs):
                        bit = sig['bit_start']
                        cond = '%s > 0u' % sig['name']
                        if i == 0:
                            L.append('    if (%s)' % cond)
                            L.append('    {')
                            L.append('        g_u8_tx_can_data[%d] = BIT%d;' % (global_byte, bit))
                            L.append('    }')
                            L.append('    else')
                            L.append('    {')
                            L.append('        g_u8_tx_can_data[%d] = 0u;' % global_byte)
                            L.append('    }')
                        else:
                            L.append('    if (%s)' % cond)
                            L.append('    {')
                            L.append('        g_u8_tx_can_data[%d] |= BIT%d;' % (global_byte, bit))
                            L.append('    }')
                elif strategy == 'shift':
                    inited = False
                    for sig in sigs:
                        if sig['bit_width'] >= 8:
                            n_bytes = sig['bit_width'] // 8
                            for j in range(n_bytes):
                                shift_val = (n_bytes - 1 - j) * 8
                                target_byte = global_byte + j
                                defined_bytes.add(target_byte)
                                if shift_val > 0:
                                    L.append('    g_u8_tx_can_data[%d] = (UINT8)(%s >> %du);'
                                             % (target_byte, sig['name'], shift_val))
                                else:
                                    L.append('    g_u8_tx_can_data[%d] = (UINT8)(%s);'
                                             % (target_byte, sig['name']))
                        elif sig['bit_width'] == 1:
                            L.append('    if (%s > 0u)' % sig['name'])
                            L.append('    {')
                            L.append('        g_u8_tx_can_data[%d] |= BIT%d;'
                                     % (global_byte, sig['bit_start']))
                            L.append('    }')
                        else:
                            if not inited:
                                L.append('    g_u8_tx_can_data[%d] = 0u;' % global_byte)
                                inited = True
                            if sig['bit_start'] == 0:
                                L.append('    g_u8_tx_can_data[%d] |= ((UINT8)%s);'
                                         % (global_byte, sig['name']))
                            else:
                                L.append('    g_u8_tx_can_data[%d] |= ((UINT8)%s << %du);'
                                         % (global_byte, sig['name'], sig['bit_start']))
            else:
                L.append('    g_u8_tx_can_data[%d] = 0u;' % global_byte)
    L.append('}')
    return '\n'.join(L)


def generate_excel_c_code(sheets):
    parts = []
    for s in sheets:
        parts.append(generate_excel_c_function(s))
        parts.append('')
    return '\n'.join(parts)


# ------------------------------------------------- ini / DBC 生成（每 Sheet 独立）

CAN_ID_RE = re.compile(r'^(?:ID|0X)?([0-9A-F]{2,3})$', re.IGNORECASE)


def _parse_can_id(raw):
    """从 F 列原始值（'7E0' / '0x7E0' / 'ID7E0'）提取大写十六进制 ID。"""
    if raw is None:
        return ''
    m = CAN_ID_RE.match(str(raw).strip())
    return m.group(1).upper() if m else ''


def _excel_frames(sheet_info):
    """按 CAN ID 出现顺序把 Sheet 内信号分帧。返回 [(can_id, [signals...]), ...]。"""
    frames = []
    index = {}
    for sig in sheet_info['signals']:
        cid = _parse_can_id(sig.get('can_id', ''))
        if cid not in index:
            index[cid] = len(frames)
            frames.append((cid, []))
        frames[index[cid]][1].append(sig)
    return frames


def _signal_ident_names(signals):
    """帧内信号名 -> 合法 DBC/ini 标识符（非法字符替换为 _，帧内去重）。"""
    used = {}
    out = []
    for sig in signals:
        base = _sanitize_c_name(sig['name'])
        if base and base[0].isdigit():
            base = '_' + base
        if base in used:
            used[base] += 1
            base = '%s_%d' % (base, used[base])
        else:
            used[base] = 0
        out.append(base)
    return out


_INI_HEADER = [
    '///配置说明：',
    '///每行以$xx yyy:起始的为有效行，$xx yyy:请勿修改',
    "///每行$xx yyy:之后的为该服务的数据，数据中以','分开",
    '///每行以///起始的为注释行，说明之后的数据存放格式，请勿删除',
    '///注意：每行除空白行、注释行和有效行之外任何数据均会报错',
    '',
    "///$Message: + Id + ','  + Length + ','  + Name + ','  + Cycle",
    '///&Format: + Intel or Motorola or Lin',
    "///@Signal: + Start + ','  + Length + ','  + Name + ',' + DefaultValue",
    "///#Value: + Value + ','  + Name",
    "///#PhyValue: + MaxValue + ','  + Factor + ',' + Offset + ',' + Unit",
    '',
]

_DBC_HEADER = [
    'VERSION " "',
    '',
    '',
    '',
    ' NS_ : ',
]
for _ns in ['NS_DESC_', 'CM_', 'BA_DEF_', 'BA_', 'VAL_', 'CAT_DEF_', 'CAT_',
            'FILTER', 'BA_DEF_DEF_', 'EV_DATA_', 'ENVVAR_DATA_', 'SGTYPE_',
            'SGTYPE_VAL_', 'BA_DEF_SGTYPE_', 'SIG_TYPE_REF_', 'SIG_GROUP_',
            'SIG_VALTYPE_', 'SIGTYPE_VALTYPE_', 'BO_TX_BU_', 'BA_DEF_REL_',
            'BA_REL_', 'BA_DEF_DEF_REL_', 'BU_SG_REL_', 'BU_EV_REL_',
            'BU_BO_REL_', 'SG_MUL_VAL_']:
    _DBC_HEADER.append('      ' + _ns)
_DBC_HEADER += [
    '',
    'BS_:',
    '',
    'BU_: BCM',
    '',
    '',
    '//monitor',
    '',
]


def generate_excel_ini_content(sheet_info):
    """为单个 Sheet 生成上位机 ini（Motorola 格式，帧 = CAN ID 分组）。"""
    L = list(_INI_HEADER)
    frames = _excel_frames(sheet_info)
    for cid, sigs in frames:
        L.append('$Message:%s,8,RX_%s,10' % (cid, cid))
        L.append('&Format:Motorola')
        names = _signal_ident_names(sigs)
        for sig, name in zip(sigs, names):
            start = sig['byte'] * 8 + sig['bit_start']
            L.append('@Signal:%d,%d,%s,0' % (start, sig['bit_width'], name))
            if sig['bit_width'] == 1:
                L.append('#Value:0,OFF;')
                L.append('#Value:1,ON;')
        L.append('')
    return '\n'.join(L)


def generate_excel_dbc_content(sheet_info):
    """为单个 Sheet 生成 DBC（消息名 RX_<ID>，BCM 发送）。"""
    L = list(_DBC_HEADER)
    msg_info = []
    for cid, sigs in _excel_frames(sheet_info):
        dec_id = int(cid, 16)
        msg_info.append((dec_id, sigs))
        L.append('BO_ %d RX_%s: 8 BCM' % (dec_id, cid))
        names = _signal_ident_names(sigs)
        for sig, name in zip(sigs, names):
            start = sig['byte'] * 8 + sig['bit_start']
            L.append('   SG_ %s: %d|%d@0+ (1,0) [0|0] ""  Vector__XXX'
                     % (name, start, sig['bit_width']))
        L.append('')
    L += [
        'BA_DEF_ BO_  "GenMsgSendType" ENUM  "Cyclic","Event";',
        'BA_DEF_ BO_  "GenMsgCycleTime" INT 0 65535;',
        'BA_DEF_  "BusType" STRING ;',
        'BA_DEF_DEF_  "GenMsgSendType" "Cyclic";',
        'BA_DEF_DEF_  "GenMsgCycleTime" 0;',
        'BA_DEF_DEF_  "BusType" "CAN";',
        'BA_ "BusType" "CAN1";',
        '',
        '',
    ]
    for dec_id, _ in msg_info:
        L.append('BA_ "GenMsgCycleTime" BO_ %d 10;' % dec_id)
    L.append('')
    for dec_id, sigs in msg_info:
        names = _signal_ident_names(sigs)
        for sig, name in zip(sigs, names):
            if sig['bit_width'] == 1:
                L.append('VAL_ %d %s 0 "OFF" 1 "ON";' % (dec_id, name))
    return '\n'.join(L)


class CanIdDialog(QDialog):
    """让用户填入各 CAN 报文的 16 进制 ID。"""

    def __init__(self, num_input_msgs, num_output_msgs, parent=None):
        super().__init__(parent)
        self.setWindowTitle('CAN ID 输入')
        self._id_edits = []

        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        total = num_input_msgs + num_output_msgs
        info = QLabel('输入 %d 个报文 + 输出 %d 个报文 = 共需 %d 个 CAN ID'
                      % (num_input_msgs, num_output_msgs, total))
        info.setStyleSheet('font-weight:bold;')
        layout.addWidget(info)

        for i in range(num_input_msgs):
            row = QHBoxLayout()
            row.addWidget(QLabel('输入报文 %d CAN ID:' % (i + 1)))
            edit = QLineEdit()
            edit.setPlaceholderText('16进制，如 7E0')
            edit.setText('%03X' % (0x7E0 + i))
            row.addWidget(edit)
            layout.addLayout(row)
            self._id_edits.append(edit)

        for i in range(num_output_msgs):
            row = QHBoxLayout()
            row.addWidget(QLabel('输出报文 %d CAN ID:' % (i + 1)))
            edit = QLineEdit()
            edit.setPlaceholderText('16进制，如 7E0')
            edit.setText('%03X' % (0x7E0 + num_input_msgs + i))
            row.addWidget(edit)
            layout.addLayout(row)
            self._id_edits.append(edit)

        btn_box = QHBoxLayout()
        btn_box.addStretch()
        ok_btn = QPushButton('确定')
        cancel_btn = QPushButton('取消')
        ok_btn.setMinimumWidth(80)
        cancel_btn.setMinimumWidth(80)
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btn_box.addWidget(ok_btn)
        btn_box.addWidget(cancel_btn)
        layout.addLayout(btn_box)

    def get_ids(self):
        return [e.text().strip() for e in self._id_edits]


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_NAME)
        self.setWindowIcon(QIcon(resource_path(LOGO_PNG)))
        self.model = None
        self.excel_model = None
        self.excel_warnings = []
        self._build_menubar()
        self._show_placeholder()
        self._style()

    # ------------------------------------------------------------------ 菜单栏
    def _build_menubar(self):
        mb = self.menuBar()

        file_menu = mb.addMenu('文件(&F)')

        import_menu = file_menu.addMenu('导入')

        act_io_enum = QAction('IO枚举文本', self)
        act_io_enum.setStatusTip('导入 I/O 枚举 txt 文件')
        act_io_enum.triggered.connect(self.on_import_io_enum)
        import_menu.addAction(act_io_enum)

        act_debug_excel = QAction('Debug表格', self)
        act_debug_excel.setStatusTip('导入 Excel 文件，自动生成 CAN 测试帧 C 函数')
        act_debug_excel.triggered.connect(self.on_import_debug_excel)
        import_menu.addAction(act_debug_excel)

        file_menu.addSeparator()

        act_exit = QAction('退出', self)
        act_exit.setShortcut(QKeySequence('Ctrl+Q'))
        act_exit.setStatusTip('退出程序')
        act_exit.triggered.connect(self.close)
        file_menu.addAction(act_exit)

        cfg_menu = mb.addMenu('配置(&C)')

        act_clear = QAction('清空', self)
        act_clear.setStatusTip('清空当前数据')
        act_clear.triggered.connect(self.on_clear)
        cfg_menu.addAction(act_clear)

        cfg_menu.addSeparator()

        act_about = QAction('关于', self)
        act_about.setStatusTip('关于本工具')
        act_about.triggered.connect(self.on_about)
        cfg_menu.addAction(act_about)

        help_menu = mb.addMenu('帮助(&H)')

        act_manual = QAction('使用教程', self)
        act_manual.setStatusTip('打开用户使用说明书（HTML）')
        act_manual.triggered.connect(self.on_help_manual)
        help_menu.addAction(act_manual)

        act_contact = QAction('联系我们', self)
        act_contact.setStatusTip('查看联系方式')
        act_contact.triggered.connect(self.on_help_contact)
        help_menu.addAction(act_contact)

    # ------------------------------------------------------------------ 界面切换
    def _clear_widget_refs(self):
        for attr in ('gen_btn', 'path_label', 'info_label', 'preview',
                     'input_table', 'output_table', 'tabs',
                     'excel_preview', 'excel_gen_btn', 'excel_path_label'):
            if hasattr(self, attr):
                setattr(self, attr, None)

    def _show_placeholder(self):
        self._clear_widget_refs()
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setAlignment(Qt.AlignCenter)
        tip = QLabel('请先导入文件')
        tip.setStyleSheet('color:#999; font-size:15px;')
        tip.setAlignment(Qt.AlignCenter)
        lay.addWidget(tip)
        self.setCentralWidget(w)
        self.resize(800, 520)

    def _build_io_enum_ui(self):
        self._clear_widget_refs()
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        info_box = QVBoxLayout()
        info_box.setSpacing(2)
        self.path_label = QLabel('未导入文件')
        self.path_label.setStyleSheet('color:#888;')
        info_box.addWidget(self.path_label)
        self.info_label = QLabel('')
        self.info_label.setStyleSheet('color:#0078d4; font-weight:bold;')
        info_box.addWidget(self.info_label)
        layout.addLayout(info_box)

        splitter = QSplitter(Qt.Horizontal)

        left = QWidget()
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 0, 0)
        self.tabs = QTabWidget()
        self.input_table = self._make_table(['序号', '输入枚举名称'])
        self.output_table = self._make_table(['序号', '输出枚举名称'])
        self.tabs.addTab(self.input_table, '输入枚举')
        self.tabs.addTab(self.output_table, '输出枚举')
        ll.addWidget(self.tabs, 1)
        splitter.addWidget(left)

        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.addWidget(QLabel('生成的函数预览'))
        self.preview = QPlainTextEdit()
        self.preview.setReadOnly(True)
        self.preview.setFont(QFont('Consolas', 10))
        rl.addWidget(self.preview, 1)
        splitter.addWidget(right)
        splitter.setSizes([520, 580])
        layout.addWidget(splitter, 1)

        bottom = QHBoxLayout()
        bottom.addStretch(1)
        self.gen_btn = QPushButton('生成')
        self.gen_btn.setMinimumHeight(34)
        self.gen_btn.setMinimumWidth(120)
        self.gen_btn.clicked.connect(self.on_generate)
        bottom.addWidget(self.gen_btn)
        layout.addLayout(bottom)

        self.resize(1100, 720)

    def _make_table(self, headers):
        t = QTableWidget(0, len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setSelectionMode(QAbstractItemView.SingleSelection)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.verticalHeader().setVisible(False)
        return t

    def _style(self):
        self.setStyleSheet('''
            QMainWindow { background:#f5f5f5; }
            QMenuBar { background:#f0f0f0; border-bottom:1px solid #ddd; }
            QMenuBar::item { padding:4px 10px; background:transparent; }
            QMenuBar::item:selected { background:#cde6f7; }
            QMenu { border:1px solid #ccc; background:white; }
            QMenu::item { padding:4px 24px 4px 20px; }
            QMenu::item:selected { background:#cde6f7; }
            QMenu::separator { height:1px; background:#ddd; margin:4px 8px; }
            QLabel { background:transparent; }
            QTableWidget { gridline-color:#e0e0e0; background:white; }
            QHeaderView::section { background:#f0f0f0; padding:4px; border:none; border-right:1px solid #ddd; border-bottom:1px solid #ddd; }
            QPlainTextEdit { border:1px solid #c8c8c8; border-radius:3px; background:#fafafa; }
        ''')
        gen_btn = getattr(self, 'gen_btn', None)
        if gen_btn is not None:
            gen_btn.setStyleSheet(
                'QPushButton { background:#0078d4; color:white; border:none; border-radius:4px; font-weight:bold; padding:6px 18px; }'
                'QPushButton:hover { background:#106ebe; }')
        excel_btn = getattr(self, 'excel_gen_btn', None)
        if excel_btn is not None:
            excel_btn.setStyleSheet(
                'QPushButton { background:#0078d4; color:white; border:none; border-radius:4px; font-weight:bold; padding:6px 18px; }'
                'QPushButton:hover { background:#106ebe; }')

    # ------------------------------------------------------------- 交互
    def on_import_io_enum(self):
        path, _ = QFileDialog.getOpenFileName(
            self, '选择 I/O 枚举 txt 文件', '',
            'Text Files (*.txt);;All Files (*.*)')
        if not path:
            return
        try:
            text = read_text_file(path)
            result = parse_io_enum(text)
        except Exception as e:
            QMessageBox.critical(self, '解析失败', str(e))
            return
        self.model = result
        self._build_io_enum_ui()
        self._style()
        self.path_label.setText(os.path.basename(path))
        self._fill_tables()
        self._update_info()
        self._render_preview()
        self.statusBar().showMessage('导入成功: %s' % path)

    def on_import_debug_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, '选择 Excel 文件', '',
            'Excel Files (*.xlsx *.xls);;All Files (*.*)')
        if not path:
            return
        if path.lower().endswith('.xls'):
            QMessageBox.warning(self, '文件格式错误',
                '暂不支持旧版 .xls 格式，请先用 Excel 将其另存为 .xlsx，\n'
                '然后重新导入（文件 → 导入 → Debug表格）。')
            return
        try:
            sheets, errors, warnings = parse_excel_file(path)
        except Exception as e:
            QMessageBox.warning(self, '文件格式错误',
                '无法读取 Excel 文件，文件可能已损坏或不是有效的 .xlsx 文件：\n%s\n\n'
                '请检查文件后重新导入（文件 → 导入 → Debug表格）。' % str(e))
            return
        if errors:
            msg = ('导入的 Excel 文件格式不满足生成规则，请修正以下问题后重新导入\n'
                   '（文件 → 导入 → Debug表格）：\n\n'
                   + '\n'.join('- ' + e for e in errors))
            if warnings:
                msg += ('\n\n行级告警：\n'
                        + '\n'.join('- ' + w for w in warnings[:10]))
            QMessageBox.warning(self, '文件格式错误', msg)
            return
        if not sheets:
            QMessageBox.warning(self, '无数据', '未在 Excel 中找到有效的工作表，请检查后重新导入。')
            return
        self.excel_model = sheets
        self.excel_warnings = warnings
        self.model = None
        self._build_debug_excel_ui()
        self._style()
        self.excel_path_label.setText(os.path.basename(path))
        self._render_excel_preview()
        self.statusBar().showMessage('导入成功: %s (%d 个函数)' % (path, len(sheets)))
        if warnings:
            shown = warnings[:15]
            more = '' if len(warnings) <= 15 else '\n... 其余 %d 条略' % (len(warnings) - 15)
            QMessageBox.information(self, '格式告警',
                '导入成功，但以下 %d 行不符合规则，已跳过：\n\n' % len(warnings)
                + '\n'.join('- ' + w for w in shown) + more)

    def _build_debug_excel_ui(self):
        self._clear_widget_refs()
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        self.excel_path_label = QLabel('未导入文件')
        self.excel_path_label.setStyleSheet('color:#888;')
        layout.addWidget(self.excel_path_label)

        info = QLabel('已解析 %d 个工作表 → 生成 %d 个 C 函数 + 每表独立 ini/DBC'
                      % (len(self.excel_model), len(self.excel_model)))
        info.setStyleSheet('color:#0078d4; font-weight:bold;')
        layout.addWidget(info)

        if self.excel_warnings:
            warn = QLabel('格式告警 %d 条（相关行已跳过）：%s%s'
                          % (len(self.excel_warnings), self.excel_warnings[0],
                             ' ...' if len(self.excel_warnings) > 1 else ''))
            warn.setStyleSheet('color:#b58900;')
            warn.setWordWrap(True)
            layout.addWidget(warn)

        layout.addWidget(QLabel('生成的 C 代码预览'))
        self.excel_preview = QPlainTextEdit()
        self.excel_preview.setReadOnly(True)
        self.excel_preview.setFont(QFont('Consolas', 10))
        layout.addWidget(self.excel_preview, 1)

        bottom = QHBoxLayout()
        bottom.addStretch(1)
        self.excel_gen_btn = QPushButton('生成')
        self.excel_gen_btn.setMinimumHeight(34)
        self.excel_gen_btn.setMinimumWidth(120)
        self.excel_gen_btn.clicked.connect(self.on_generate_excel)
        bottom.addWidget(self.excel_gen_btn)
        layout.addLayout(bottom)

        self.resize(1100, 720)

    def _render_excel_preview(self):
        code = generate_excel_c_code(self.excel_model)
        self.excel_preview.setPlainText(code)

    def on_generate_excel(self):
        code = generate_excel_c_code(self.excel_model)
        path, _ = QFileDialog.getSaveFileName(
            self, '保存 C 文件', 'app_can_test_debug',
            'C Source (*.c);;All Files (*.*)')
        if not path:
            return
        out_dir = os.path.dirname(path) or '.'
        written = [path]
        try:
            with open(path, 'w', encoding='utf-8', newline='\n') as f:
                f.write(code)
            for sheet in self.excel_model:
                base = _sanitize_c_name(sheet['name'])
                ini_path = os.path.join(out_dir, base + '.ini')
                dbc_path = os.path.join(out_dir, base + '.dbc')
                with open(ini_path, 'w', encoding='utf-8', newline='\r\n') as f:
                    f.write(generate_excel_ini_content(sheet))
                with open(dbc_path, 'w', encoding='utf-8', newline='\r\n') as f:
                    f.write(generate_excel_dbc_content(sheet))
                written += [ini_path, dbc_path]
        except OSError as e:
            QMessageBox.critical(self, '保存失败', str(e))
            return
        self.statusBar().showMessage('已生成 %d 个文件' % len(written))
        QMessageBox.information(
            self, '完成',
            '已生成 %d 个文件：\n%s' % (len(written), '\n'.join(written)))

    def on_clear(self):
        if self.model is None and self.excel_model is None:
            return
        self.model = None
        self.excel_model = None
        self.excel_warnings = []
        self._show_placeholder()
        self._style()
        self.statusBar().showMessage('已清空')

    def on_about(self):
        QMessageBox.about(self, '关于',
            '<div style="font-size:18px;"><b>%s</b></div>'
            '<div style="color:#808080; font-size:11px;">版本号：%s</div><br>'
            '<b>IO枚举文本</b><br>'
            '导入 I/O 枚举 txt 文件，自动识别 NUM_OF_INPUTS / NUM_OF_OUTPUTS，'
            '生成 .c 函数与上位机 .ini / .dbc 文件。<br><br>'
            '<b>Debug表格</b><br>'
            '导入 Excel 文件（跳过第 1 个工作表，为每个工作表生成 CAN 测试帧 C 函数，'
            '并为每个工作表独立生成上位机 .ini 与 .dbc 文件。'
            % (APP_NAME, APP_VERSION))

    def on_help_manual(self):
        src = resource_path(MANUAL_HTML)
        if not os.path.exists(src):
            QMessageBox.warning(self, '错误', '未找到使用教程文件：%s' % MANUAL_HTML)
            return
        out_dir = os.path.join(os.environ.get('TEMP', os.path.expanduser('~')),
                               APP_NAME)
        dst = os.path.join(out_dir, MANUAL_HTML)
        try:
            os.makedirs(out_dir, exist_ok=True)
            if (not os.path.exists(dst)
                    or os.path.getsize(dst) != os.path.getsize(src)):
                shutil.copyfile(src, dst)
        except OSError as e:
            QMessageBox.critical(self, '错误', '无法准备使用教程文件：\n%s' % e)
            return
        os.startfile(dst)

    def on_help_contact(self):
        QMessageBox.information(self, '联系我们', '联系邮箱：1815393632@qq.com')

    def _fill_tables(self):
        r = self.model
        self._fill_table(self.input_table, r['inputs'])
        self._fill_table(self.output_table, r['outputs'])
        self.tabs.setTabText(0, '输入枚举 (%d)' % len(r['inputs']))
        self.tabs.setTabText(1, '输出枚举 (%d)' % len(r['outputs']))

    def _fill_table(self, table, names):
        table.setRowCount(len(names))
        for i, n in enumerate(names):
            table.setItem(i, 0, QTableWidgetItem(str(i)))
            table.setItem(i, 1, QTableWidgetItem(n))

    def _update_info(self):
        r = self.model
        self.info_label.setText(
            'NUM_OF_INPUTS = %d    |    NUM_OF_OUTPUTS = %d'
            % (r['num_inputs'], r['num_outputs']))

    def _render_preview(self):
        r = self.model
        self.preview.setPlainText(
            generate_function_text(r['num_inputs'], r['num_outputs']))

    def on_generate(self):
        if not self.model:
            return
        r = self.model
        num_in_msgs = math.ceil(r['num_inputs'] / 64)
        num_out_msgs = math.ceil(r['num_outputs'] / 64)

        dlg = CanIdDialog(num_in_msgs, num_out_msgs, self)
        if dlg.exec_() != QDialog.Accepted:
            return
        can_ids = dlg.get_ids()
        if any(not eid for eid in can_ids):
            QMessageBox.warning(self, '提示', 'CAN ID 不能为空，请重新填写。')
            return

        c_content = generate_full_c_content(r['num_inputs'], r['num_outputs'])
        ini_content = generate_ini_content(r['inputs'], r['outputs'], can_ids)
        dbc_content = generate_dbc_content(r['inputs'], r['outputs'], can_ids)

        path, _ = QFileDialog.getSaveFileName(
            self, '保存文件', 'app_can_test_frame_01_input_output',
            'C Source (*.c);;All Files (*.*)')
        if not path:
            return

        base = os.path.splitext(path)[0]
        ini_path = base + '.ini'
        dbc_path = base + '.dbc'

        try:
            with open(path, 'w', encoding='utf-8', newline='\n') as f:
                f.write(c_content)
            with open(ini_path, 'w', encoding='utf-8', newline='\r\n') as f:
                f.write(ini_content)
            with open(dbc_path, 'w', encoding='utf-8', newline='\r\n') as f:
                f.write(dbc_content)
        except OSError as e:
            QMessageBox.critical(self, '保存失败', str(e))
            return

        self.statusBar().showMessage('已生成: %s + %s + %s' % (path, ini_path, dbc_path))
        QMessageBox.information(self, '完成',
            '已生成：\n%s\n%s\n%s' % (path, ini_path, dbc_path))


def selftest(path):
    text = read_text_file(path)
    r = parse_io_enum(text)
    print('NUM_OF_INPUTS  =', r['num_inputs'])
    print('NUM_OF_OUTPUTS =', r['num_outputs'])
    num_in = math.ceil(r['num_inputs'] / 64)
    num_out = math.ceil(r['num_outputs'] / 64)
    ids = ['%03X' % (0x7E0 + i) for i in range(num_in)]
    ids += ['%03X' % (0x7E0 + num_in + i) for i in range(num_out)]
    print('CAN IDs needed:', ids)
    print('---- generated .ini (first 30 lines) ----')
    ini = generate_ini_content(r['inputs'], r['outputs'], ids).split('\n')
    print('\n'.join(ini[:30]))


def main():
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(resource_path(LOGO_PNG)))
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    if len(sys.argv) >= 3 and sys.argv[1] == '--selftest':
        selftest(sys.argv[2])
    else:
        main()
